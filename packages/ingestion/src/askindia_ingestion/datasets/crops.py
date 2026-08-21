"""DA&FW five-year estimates of area, production and yield of principal crops by state and season.

Source: Directorate of Economics & Statistics, Department of Agriculture & Farmers Welfare
(DA&FW) — the "5-Year Estimates of Foodgrains, Oilseeds and other Commercial Crops" workbook
linked from https://desagri.gov.in/statistics-type/five-year-estimates/. One sheet per crop
(33 at the time of writing), each a wide table: Crop | State | Season | Area x 5 years |
Production x 5 years | Yield x 5 years, with 3-5 title rows above a two-row header, the crop
name printed only on the first data row (or not at all) and the state name only on its first
season row. The upload path of the xlsx changes with every release, so fetch_raw discovers it
from the listing page instead of hardcoding it.
"""

from __future__ import annotations

import io
import logging
import re
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import openpyxl
import pandas as pd
from bs4 import BeautifulSoup

from askindia_ingestion.contracts import (
    BaseLoader,
    Cadence,
    ColumnSpec,
    DatasetSpec,
    RawArtifact,
    SourceFormat,
)
from askindia_ingestion.loaders.http import fetch

LISTING_URL = "https://desagri.gov.in/statistics-type/five-year-estimates/"
SEASONS = ("Kharif", "Rabi", "Summer", "Total")
MEASURES = {
    "area": "area_thousand_ha",
    "production": "production_thousand_tonnes",
    "yield": "yield_kg_per_ha",
}
log = logging.getLogger(__name__)

Row = tuple[Any, ...]

_TITLE = re.compile(r"Estimates?\s+of\s+Area,?\s*Production\s*&\s*Yield\s+for\s+(.+?)\s*$", re.I)
# "Data for the year 2025-26 is of 3ʳᵈ Advance Estimates" — the ordinal suffix is superscript.
_ADVANCE = re.compile(
    r"Data\s+for\s+the\s+year\s+(\d{4}-\d{2})\s+is\s+of\s+(\d+)\D{0,6}?Advance\s+Estimate", re.I
)
# "# Cotton Production in Thousand Bales, 1Bale=170 Kg"
_BALES = re.compile(
    r"Production\s+in\s+Thousand\s+Bales,?\s*1\s*Bale\s*=\s*(\d+(?:\.\d+)?)\s*Kg", re.I
)
_YEAR = re.compile(r"^\d{4}-\d{2}$")
_NULL_TEXT = {"", "-", "--", "NA", "N/A", "NR"}

SPEC = DatasetSpec(
    key="crop_production",
    title="Area, production and yield of principal crops by state and season (DA&FW estimates)",
    source_org=(
        "Directorate of Economics & Statistics, Department of Agriculture & Farmers Welfare, "
        "Ministry of Agriculture & Farmers Welfare"
    ),
    source_url=LISTING_URL,
    landing_url=LISTING_URL,
    fmt=SourceFormat.XLSX,
    cadence=Cadence.ANNUAL,
    table_name="data.crop_production",
    columns=(
        ColumnSpec(
            name="crop",
            pg_type="text",
            description=(
                "Crop or crop group as titled on its sheet, e.g. 'Rice', 'Wheat', 'Total "
                "Food Grains', 'Nutri/Coarse Cereals', 'Total Oil Seeds', 'Jute & Mesta'"
            ),
        ),
        ColumnSpec(
            name="state",
            pg_type="text",
            description=(
                "State or union territory as published, 'All India' for the national total, "
                "'Others' for small states/UTs grouped by the source"
            ),
        ),
        ColumnSpec(
            name="season",
            pg_type="text",
            description=(
                "Crop season: Kharif (monsoon-sown), Rabi (winter-sown), Summer, or 'Total' "
                "for the whole crop year"
            ),
            allowed=SEASONS,
        ),
        ColumnSpec(
            name="crop_year",
            pg_type="text",
            description="Agricultural year July-June, written '2021-22' for July 2021 to June 2022",
        ),
        ColumnSpec(
            name="estimate_type",
            pg_type="text",
            description=(
                "'final' for years published as Final Estimates, or e.g. '3rd advance "
                "estimate' where the sheet says that year is an Advance Estimate"
            ),
        ),
        ColumnSpec(
            name="area_thousand_ha",
            pg_type="numeric(12,2)",
            description="Area sown under the crop",
            unit="thousand hectares",
            min=0,
            max=1_000_000,
            nullable=True,
        ),
        ColumnSpec(
            name="production_thousand_tonnes",
            pg_type="numeric(14,2)",
            description=(
                "Production; for cotton, jute and mesta the source publishes thousand bales and "
                "this column is bales x kg per bale / 1000 using the factor printed on the sheet"
            ),
            unit="thousand tonnes",
            min=0,
            max=2_000_000,
            nullable=True,
        ),
        ColumnSpec(
            name="yield_kg_per_ha",
            pg_type="numeric(10,2)",
            description="Yield as published (production per hectare of area sown)",
            unit="kg per hectare",
            min=0,
            max=500_000,
            nullable=True,
        ),
    ),
    unique_key=("crop", "state", "season", "crop_year"),
    min_rows=5700,
    caveats=(
        "Earlier crop years are DA&FW Final Estimates; the latest crop year is an Advance "
        "Estimate (estimate_type says which, e.g. '3rd advance estimate') and is revised in "
        "later releases, so compare like with like when checking a year-on-year claim.",
        "crop_year is the agricultural year July to June: '2021-22' runs July 2021 to June 2022. "
        "Kharif is the monsoon crop, Rabi the winter crop; 'Total' is the whole-year figure "
        "published by the source, not recomputed here.",
        "Units: area in thousand hectares, production in thousand tonnes, yield in kg per "
        "hectare, all as published. Multiply thousand tonnes by 0.001 for million tonnes.",
        "Cotton production is published in thousand bales of 170 kg and jute/mesta in thousand "
        "bales of 180 kg; production_thousand_tonnes converts those with the factor printed on "
        "the sheet (bales x kg per bale / 1000). yield_kg_per_ha for these crops is lint/fibre "
        "per hectare as published.",
        "'All India' is a state value carrying the national total; do not sum states to get it. "
        "State names are as published and vary between sheets ('A&N Islands' vs 'Andaman And "
        "Nicobar Islands', 'J&K' vs 'Jammu And Kashmir', 'DNH' vs 'Dadra And Nagar Haveli'); "
        "'Others' groups small states/UTs.",
        "Crop groups (Total Food Grains, Total Pulses, Total Oil Seeds, Nutri/Coarse Cereals, "
        "Shree Anna /Nutri Cereals, Jute & Mesta) are separate crop values alongside their "
        "components; never add crops together.",
        "A crop x state x season x year row is omitted when the sheet leaves area, production "
        "and yield all blank (not grown, or not yet estimated); a single blank cell is NULL.",
        "Only the five crop years of the current release are covered; earlier history is not "
        "in this dataset.",
        "desagri.gov.in serves an incomplete certificate chain, so the listing page and workbook "
        "are fetched without TLS verification; the file hash is recorded on every load.",
    ),
    difficulty="medium",
    verify_tls=False,
)


def _text(cell: Any) -> str:
    return "" if cell is None else re.sub(r"\s+", " ", str(cell)).strip()


def _num(cell: Any) -> float | None:
    if cell is None or isinstance(cell, bool):
        if cell is None:
            return None
        raise ValueError(f"boolean cell {cell!r}")
    if isinstance(cell, int | float):
        return float(cell)
    text = _text(cell).replace(",", "")
    if text.upper() in _NULL_TEXT:
        return None
    return float(text)


def _ordinal(n: int) -> str:
    suffix = "th" if 10 <= n % 100 <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _header_index(rows: Sequence[Row], sheet: str) -> int:
    for i, row in enumerate(rows):
        if len(row) >= 3 and [_text(c).lower() for c in row[:3]] == ["crop", "state", "season"]:
            if i == 0:
                raise ValueError(f"{sheet}: header is the first row; expected a title row above")
            return i
    raise ValueError(f"{sheet}: no 'Crop | State | Season' header row found")


def _column_map(span_row: Row, year_row: Row, sheet: str) -> list[tuple[int, str, str]]:
    """(column index, output column, crop year) for every year column, from the two header rows.

    The row above the years carries 'Area', 'Production', 'Yield' once each at the start of its
    block (merged cells), so the measure is forward-filled across the years to its right.
    """
    out: list[tuple[int, str, str]] = []
    measure: str | None = None
    for idx in range(3, len(year_row)):
        label = _text(span_row[idx]).lower() if idx < len(span_row) else ""
        if label:
            if label not in MEASURES:
                raise ValueError(f"{sheet}: unexpected measure header {label!r}")
            measure = MEASURES[label]
        year = _text(year_row[idx])
        if not year:
            continue
        if not _YEAR.match(year):
            raise ValueError(f"{sheet}: unexpected year header {year!r}")
        if measure is None:
            raise ValueError(f"{sheet}: year column {year!r} has no Area/Production/Yield heading")
        out.append((idx, measure, year))
    found = {m for _, m, _ in out}
    if found != set(MEASURES.values()):
        raise ValueError(f"{sheet}: expected Area, Production and Yield blocks, found {found}")
    return out


def _crop_name(rows: Sequence[Row], hdr: int, sheet: str) -> str:
    for row in rows[:hdr]:
        match = _TITLE.search(_text(row[0]) if row else "")
        if match:
            return match.group(1)
    log.warning("%s: no title row names the crop; using the sheet name", sheet)
    return sheet


def _footnotes(rows: Sequence[Row], hdr: int) -> list[str]:
    """Text-only rows below the header: first cell filled, state and season empty."""
    return [
        _text(row[0])
        for row in rows[hdr + 1 :]
        if len(row) >= 3 and _text(row[0]) and not _text(row[1]) and not _text(row[2])
    ]


def advance_years(notes: Sequence[str]) -> dict[str, str]:
    """crop year -> estimate label, from footnotes like 'Data for the year 2025-26 is of 3rd AE'."""
    out: dict[str, str] = {}
    for note in notes:
        match = _ADVANCE.search(note)
        if match:
            year, nth = match.groups()
            out[year] = f"{_ordinal(int(nth))} advance estimate"
    return out


def _kg_per_bale(notes: Sequence[str]) -> float | None:
    for note in notes:
        match = _BALES.search(note)
        if match:
            return float(match.group(1))
    return None


def parse_sheet(rows: Sequence[Row], sheet: str, advance: Mapping[str, str]) -> pd.DataFrame:
    """One crop sheet -> long frame, one row per state x season x crop year with any data."""
    hdr = _header_index(rows, sheet)
    columns = _column_map(rows[hdr - 1], rows[hdr], sheet)
    crop = _crop_name(rows, hdr, sheet)
    notes = _footnotes(rows, hdr)
    kg_per_bale = _kg_per_bale(notes)
    years = sorted({year for _, _, year in columns})
    records: list[dict[str, object]] = []
    state = ""
    for row in rows[hdr + 1 :]:
        if len(row) < 3:
            continue
        if _text(row[1]):
            state = _text(row[1])
        season = _text(row[2])
        if not season:
            continue
        if not state:
            raise ValueError(f"{sheet}: season row {season!r} appears before any state name")
        values: dict[str, dict[str, float | None]] = {year: {} for year in years}
        for idx, measure, year in columns:
            cell = row[idx] if idx < len(row) else None
            try:
                values[year][measure] = _num(cell)
            except ValueError as e:
                raise ValueError(
                    f"{sheet}: non-numeric cell {cell!r} for {state}/{season}/{year} ({e})"
                ) from None
        for year in years:
            measures = values[year]
            if all(v is None for v in measures.values()):
                continue
            production = measures.get("production_thousand_tonnes")
            if production is not None and kg_per_bale is not None:
                production = round(production * kg_per_bale / 1000, 2)
            records.append(
                {
                    "crop": crop,
                    "state": state,
                    "season": season,
                    "crop_year": year,
                    "estimate_type": advance.get(year, "final"),
                    "area_thousand_ha": measures.get("area_thousand_ha"),
                    "production_thousand_tonnes": production,
                    "yield_kg_per_ha": measures.get("yield_kg_per_ha"),
                }
            )
    if not records:
        raise ValueError(f"{sheet}: no data rows parsed")
    return pd.DataFrame.from_records(records)


def parse_workbook(content: bytes, source: str = "") -> pd.DataFrame:
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheets: dict[str, list[Row]] = {
            name: [tuple(row) for row in workbook[name].iter_rows(values_only=True)]
            for name in workbook.sheetnames
        }
    finally:
        workbook.close()
    if not sheets:
        raise ValueError("workbook has no sheets")

    # Which crop years are advance estimates is stated in a footnote on (nearly) every sheet;
    # sheets that omit it inherit the workbook-wide answer, and disagreement fails loud.
    advance: dict[str, str] = {}
    for name, rows in sheets.items():
        for year, label in advance_years(_footnotes(rows, _header_index(rows, name))).items():
            if advance.setdefault(year, label) != label:
                raise ValueError(
                    f"{name}: {year} is {label!r} but another sheet says {advance[year]!r}"
                )
    if not advance and re.search(r"\bAE\b|advance", source, re.I):
        raise ValueError(
            "the source is labelled an advance estimate but no sheet says which crop year is one"
        )

    frame = pd.concat(
        [parse_sheet(rows, name, advance) for name, rows in sheets.items()], ignore_index=True
    )
    for column in MEASURES.values():
        frame[column] = pd.to_numeric(frame[column])
    return frame.reset_index(drop=True)


def _slug(url: str) -> str:
    """Filename of a URL reduced to lowercase alphanumerics, for loose matching."""
    return re.sub(r"[^a-z0-9]", "", url.rsplit("/", 1)[-1].lower())


def discover_xlsx_url(html: str, base_url: str = LISTING_URL) -> str:
    """The five-year estimates workbook link on the listing page (newest is listed first)."""
    soup = BeautifulSoup(html, "lxml")
    links: list[str] = []
    for anchor in soup.select("a[href]"):
        href = urljoin(base_url, str(anchor.get("href") or "").strip())
        if href.split("?", 1)[0].lower().endswith(".xlsx") and href not in links:
            links.append(href)
    if not links:
        raise ValueError(f"no .xlsx link found on {base_url}")
    five_year = [url for url in links if "fiveyear" in _slug(url)]
    if len(five_year) > 1:
        log.warning(
            "%d five-year workbooks listed; taking the first: %s", len(five_year), five_year
        )
    if five_year:
        return five_year[0]
    if len(links) == 1:
        return links[0]
    raise ValueError(f"{len(links)} .xlsx links on {base_url} and none is a five-year workbook")


class CropProductionLoader(BaseLoader):
    def fetch_raw(self) -> RawArtifact:
        listing = fetch(self.spec.key, LISTING_URL, verify_tls=self.spec.verify_tls)
        url = discover_xlsx_url(listing.content.decode("utf-8", errors="replace"), listing.url)
        log.info("%s: workbook discovered at %s", self.spec.key, url)
        return fetch(self.spec.key, url, verify_tls=self.spec.verify_tls)

    def parse(self, raw: RawArtifact) -> pd.DataFrame:
        return parse_workbook(raw.content, source=raw.url)


def build(snapshot_dir: Path | None = None) -> BaseLoader:
    return CropProductionLoader(SPEC, snapshot_dir=snapshot_dir)
